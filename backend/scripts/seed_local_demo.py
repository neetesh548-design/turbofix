"""Make the checked-in local workbook useful for a complete post-login demo.

Run from the repository root with:
    backend/.venv/bin/python backend/scripts/seed_local_demo.py
"""

from datetime import datetime, timezone
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[2]
WORKBOOK = ROOT / "backend" / "TurboFix-Tracker.xlsx"


def row_map(ws):
    headers = [cell.value for cell in ws[1]]
    return headers, {row[0].value: row[0].row for row in ws.iter_rows(min_row=2) if row[0].value}


def set_value(ws, headers, row_number, field, value):
    ws.cell(row=row_number, column=headers.index(field) + 1, value=value)


def main():
    workbook = openpyxl.load_workbook(WORKBOOK)

    tickets = workbook["Tickets"]
    ticket_headers, ticket_rows = row_map(tickets)
    corrections = {
        "T001": {"company_code": "ACME3", "machine_name": "CNC Lathe 1"},
        "T002": {"company_code": "ACME3", "machine_name": "Hydraulic Press"},
        "T003": {"company_code": "BETA1", "machine_name": "Grinding Machine"},
        "T004": {"company_code": "BETA1", "machine_name": "Compressor"},
    }
    for ticket_id, values in corrections.items():
        if ticket_id in ticket_rows:
            for field, value in values.items():
                set_value(tickets, ticket_headers, ticket_rows[ticket_id], field, value)

    demo_tickets = [
        # 5 High-Priority Demo Tickets
        {
            "ticket_id": "T005", "machine_id": "TF-ACME3-M001", "company_code": "ACME3",
            "machine_name": "CNC Lathe 1", "reported_at": datetime(2026, 7, 10, 11, 15),
            "reporter_phone": "+919900011115", "description": "Spindle vibration returned after restart",
            "ai_summary": "Repeat spindle bearing issue; inspect bearing seating and lubrication circuit",
            "urgency": "High", "status": "reported", "language": "en",
        },
        {
            "ticket_id": "T006", "machine_id": "TF-ACME3-M002", "company_code": "ACME3",
            "machine_name": "Hydraulic Press", "reported_at": datetime(2026, 7, 12, 14, 20),
            "reporter_phone": "+919900011116", "description": "Pressure drops during the final stroke",
            "ai_summary": "Hydraulic pressure loss; check seal kit and pressure relief valve",
            "urgency": "High", "status": "reported", "language": "en",
        },
        {
            "ticket_id": "T007", "machine_id": "TF-ACME3-M003", "company_code": "ACME3",
            "machine_name": "Grinding Machine", "reported_at": datetime(2026, 7, 13, 9, 45),
            "reporter_phone": "+919900011117", "description": "Coolant pump failed - no fluid circulation",
            "ai_summary": "Pump motor seized; check impeller blockage and motor winding resistance",
            "urgency": "High", "status": "reported", "language": "en",
        },
        {
            "ticket_id": "T008", "machine_id": "TF-ACME3-M004", "company_code": "ACME3",
            "machine_name": "Compressor Unit", "reported_at": datetime(2026, 7, 14, 16, 30),
            "reporter_phone": "+919900011118", "description": "Air discharge temperature exceeds 95°C",
            "ai_summary": "Intercooler fouled or thermal valve stuck; inspect cooling passages and replace if necessary",
            "urgency": "High", "status": "reported", "language": "en",
        },
        {
            "ticket_id": "T009", "machine_id": "TF-ACME3-M005", "company_code": "ACME3",
            "machine_name": "Packaging Machine", "reported_at": datetime(2026, 7, 15, 13, 10),
            "reporter_phone": "+919900011119", "description": "Servo motor position drift - wrapping quality degraded",
            "ai_summary": "Encoder misalignment or servo calibration drift; recalibrate position feedback and motor parameters",
            "urgency": "High", "status": "reported", "language": "en",
        },
    ]
    for demo_ticket in demo_tickets:
        if demo_ticket["ticket_id"] not in ticket_rows:
            tickets.append([demo_ticket.get(field, "") for field in ticket_headers])

    machines = workbook["Machines"]
    machine_headers, machine_rows = row_map(machines)
    open_machine_ids = {"TF-ACME3-M001", "TF-ACME3-M002"}
    for machine_id, row_number in machine_rows.items():
        set_value(machines, machine_headers, row_number, "has_open_tickets", machine_id in open_machine_ids)

    events = workbook["MachineEvents"]
    event_headers, event_rows = row_map(events)
    if "EVT-DEMO-001" not in event_rows:
        events.append([
            "EVT-DEMO-001", "TF-ACME3-M001", "ACME3", "T005", "breakdown",
            datetime(2026, 7, 10, 11, 15), "+919900011115", "Spindle vibration returned after restart",
            "text", "", "en",
        ])

    workbook.save(WORKBOOK)
    print(f"Seeded local demo workbook: {WORKBOOK}")
    print("Login: rakesh@acmeforge.example / AcmeOwner@2026")


if __name__ == "__main__":
    main()
