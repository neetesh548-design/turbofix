"""Excel-backed technician work records."""

import threading
from typing import List, Optional

import openpyxl

from app.repositories.base import TECHNICIAN_WORK_HEADER, TechnicianWorkRepository


class LocalTechnicianWorkRepository(TechnicianWorkRepository):
    def __init__(self, xlsx_path: str):
        self._path = xlsx_path
        self._lock = threading.Lock()

    @staticmethod
    def _ensure_sheet(workbook):
        if "TechnicianWork" not in workbook.sheetnames:
            worksheet = workbook.create_sheet("TechnicianWork")
            worksheet.append(TECHNICIAN_WORK_HEADER)

    def get(self, ticket_id: str) -> Optional[dict]:
        workbook = openpyxl.load_workbook(self._path, data_only=True)
        if "TechnicianWork" not in workbook.sheetnames:
            return None
        for values in workbook["TechnicianWork"].iter_rows(min_row=2, values_only=True):
            if values and values[0] == ticket_id:
                padded = list(values) + [""] * (len(TECHNICIAN_WORK_HEADER) - len(values))
                return dict(zip(TECHNICIAN_WORK_HEADER, padded))
        return None

    def list_company(self, company_code: str) -> List[dict]:
        workbook = openpyxl.load_workbook(self._path, data_only=True)
        if "TechnicianWork" not in workbook.sheetnames:
            return []
        records = []
        for values in workbook["TechnicianWork"].iter_rows(min_row=2, values_only=True):
            if not values or values[1] != company_code:
                continue
            padded = list(values) + [""] * (len(TECHNICIAN_WORK_HEADER) - len(values))
            records.append(dict(zip(TECHNICIAN_WORK_HEADER, padded)))
        return records

    def upsert(self, row: dict) -> None:
        with self._lock:
            workbook = openpyxl.load_workbook(self._path)
            self._ensure_sheet(workbook)
            worksheet = workbook["TechnicianWork"]
            for row_number in range(2, worksheet.max_row + 1):
                if worksheet.cell(row=row_number, column=1).value == row["ticket_id"]:
                    for column, key in enumerate(TECHNICIAN_WORK_HEADER, start=1):
                        worksheet.cell(row=row_number, column=column).value = row.get(key, "")
                    workbook.save(self._path)
                    return
            worksheet.append([row.get(key, "") for key in TECHNICIAN_WORK_HEADER])
            workbook.save(self._path)
