"""Local Excel implementation of the AI machine-record repository."""

import threading
from typing import List, Optional

import openpyxl

from app.repositories.base import (
    MACHINE_RECORDS_HEADER,
    MachineRecordRepository,
    new_machine_record_id,
)


class LocalMachineRecordRepository(MachineRecordRepository):
    def __init__(self, xlsx_path: str):
        self._path = xlsx_path
        self._lock = threading.Lock()

    def _ensure_sheet(self, workbook):
        if "AIRecords" not in workbook.sheetnames:
            worksheet = workbook.create_sheet("AIRecords")
            worksheet.append(MACHINE_RECORDS_HEADER)

    def next_record_id(self) -> str:
        return new_machine_record_id()

    def list(
        self,
        company_code: str,
        machine_id: Optional[str] = None,
        status: Optional[str] = None,
    ) -> List[dict]:
        workbook = openpyxl.load_workbook(self._path, data_only=True)
        if "AIRecords" not in workbook.sheetnames:
            return []
        records = []
        for row in workbook["AIRecords"].iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            record = dict(zip(MACHINE_RECORDS_HEADER, row))
            if record.get("company_code") != company_code:
                continue
            if machine_id is not None and record.get("machine_id") != machine_id:
                continue
            if status is not None and record.get("status") != status:
                continue
            records.append(record)
        return records

    def get(self, record_id: str) -> Optional[dict]:
        workbook = openpyxl.load_workbook(self._path, data_only=True)
        if "AIRecords" not in workbook.sheetnames:
            return None
        for row in workbook["AIRecords"].iter_rows(min_row=2, values_only=True):
            if row and row[0] == record_id:
                return dict(zip(MACHINE_RECORDS_HEADER, row))
        return None

    def add(self, row: dict) -> None:
        with self._lock:
            workbook = openpyxl.load_workbook(self._path)
            self._ensure_sheet(workbook)
            workbook["AIRecords"].append(
                [row.get(column, "") for column in MACHINE_RECORDS_HEADER]
            )
            workbook.save(self._path)

    def update(self, record_id: str, updates: dict) -> bool:
        with self._lock:
            workbook = openpyxl.load_workbook(self._path)
            if "AIRecords" not in workbook.sheetnames:
                return False
            worksheet = workbook["AIRecords"]
            for row_number in range(2, worksheet.max_row + 1):
                if worksheet.cell(row=row_number, column=1).value != record_id:
                    continue
                for key, value in updates.items():
                    if key in MACHINE_RECORDS_HEADER:
                        worksheet.cell(
                            row=row_number,
                            column=MACHINE_RECORDS_HEADER.index(key) + 1,
                        ).value = value
                workbook.save(self._path)
                return True
        return False

    def find_by_hash(
        self, company_code: str, machine_id: str, file_hash: str
    ) -> Optional[dict]:
        for record in self.list(company_code, machine_id):
            if record.get("file_hash") == file_hash:
                return record
        return None
