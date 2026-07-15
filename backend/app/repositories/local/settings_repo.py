"""Excel-backed company settings repository for local development and tests."""

import threading
from typing import Optional

import openpyxl

from app.repositories.base import SETTINGS_HEADER, SettingsRepository


class LocalSettingsRepository(SettingsRepository):
    def __init__(self, xlsx_path: str):
        self._path = xlsx_path
        self._lock = threading.Lock()

    @staticmethod
    def _ensure_sheet(workbook):
        if "Settings" not in workbook.sheetnames:
            worksheet = workbook.create_sheet("Settings")
            worksheet.append(SETTINGS_HEADER)

    def get(self, company_code: str) -> Optional[dict]:
        workbook = openpyxl.load_workbook(self._path, data_only=True)
        if "Settings" not in workbook.sheetnames:
            return None
        for values in workbook["Settings"].iter_rows(min_row=2, values_only=True):
            if values and values[0] == company_code:
                padded = list(values) + [""] * (len(SETTINGS_HEADER) - len(values))
                return dict(zip(SETTINGS_HEADER, padded))
        return None

    def upsert(self, row: dict) -> None:
        with self._lock:
            workbook = openpyxl.load_workbook(self._path)
            self._ensure_sheet(workbook)
            worksheet = workbook["Settings"]
            for row_number in range(2, worksheet.max_row + 1):
                if worksheet.cell(row=row_number, column=1).value == row["company_code"]:
                    for column, key in enumerate(SETTINGS_HEADER, start=1):
                        worksheet.cell(row=row_number, column=column).value = row.get(key, "")
                    workbook.save(self._path)
                    return
            worksheet.append([row.get(key, "") for key in SETTINGS_HEADER])
            workbook.save(self._path)
